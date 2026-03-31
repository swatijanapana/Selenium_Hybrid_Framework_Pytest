import openpyxl

def get_login_data():
   wb = openpyxl.load_workbook("TestData/login_data.xlsx")
   ws = wb["LoginData"]

   data = []

   for row in ws.iter_rows(min_row=2, values_only=True):
      row = list(row)

      while len(row) < 5:
         row.append("")

      data.append(tuple(row[:5])) # row = (scenario, username, password, expected,expected_message)
   return  data


def get_api_test_data():
   wb = openpyxl.load_workbook("TestData/test_data.xlsx")
   ws = wb["APIUsers"]

   data = []

   headers = [cell.value for cell in ws[1]]

   for row in ws.iter_rows(min_row=2, values_only=True):
       row_data = dict(zip(headers, row))
       data.append(row_data)

   return  data